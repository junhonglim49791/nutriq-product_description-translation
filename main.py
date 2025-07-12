import pandas as pd
from openpyxl import load_workbook
import time
import asyncio
from bs4 import BeautifulSoup
from googletrans import Translator
import re


async def translate_text_in_bulk(text_list):
    async with Translator(timeout=20.0) as translator:
        translated_objects = await translator.translate(text_list, src="en", dest="ms")
        translated_text = []
        for translated_object in translated_objects:
            translated_text.append(translated_object.text)
        return translated_text


async def translate_text(text):
    async with Translator(timeout=10.0) as translator:
        translate_obj = await translator.translate(text, src="en", dest="ms")
        return translate_obj.text


original_file_name = "basic300734464062export1752332755069_0712-23-05-55.xlsx"

df = pd.read_excel(
    original_file_name,
    header=0,
    skiprows=[1, 2, 3],
)

# save to excel file with the original format and structure, else all will be "squeeze" in to the cells
wb = load_workbook(original_file_name)
ws = wb.active

start_time = time.time()

translated_html_string_list = []
for i in range(0, df.shape[0]):

    # Skip empty rows
    if pd.isna(df.loc[i, "Main Description"]):
        translated_html_string_list.append("")
        continue

    soup = BeautifulSoup(df.loc[i, "Main Description"], "html.parser")
    print(f"Current row index: {i}")

    for text in soup.find_all(string=True):
        try:
            if text.strip():
                print(f"Current text: {text}")
                # Ingredients come with alot of "," that affects translations accurcy, send them in chunks to translate
                if text.count(",") > 10:
                    chunks = text.split(",")
                    translated_chunks_list = []
                    for chunk in chunks:
                        translated_chunk = asyncio.run(translate_text(chunk))
                        translated_chunks_list.append(translated_chunk)
                    translated_text = ", ".join(translated_chunks_list)
                else:
                    translated_text = asyncio.run(translate_text(text))

                # Prevent kata ganda to have spaces. Example: kanak -kanak
                # But this also makes - text become -text. Its basically removing the spaces before and after "-"
                translated_text = re.sub(r"\s*([\-])\s*", r"\1", translated_text)
                text.replace_with(translated_text)
                print(f"Translated text: {translated_text}")

        except Exception as e:
            print(f"Error Message: {e}")

    translated_html_string_list.append(str(soup))

# get_loc() starts from 0, where ws.cell() starts from 1
main_description_index = df.columns.get_loc("Main Description") + 1

# replace english description to malay description
for i, val in enumerate(translated_html_string_list, start=5):
    ws.cell(row=i, column=main_description_index).value = val

wb.save("malay_description.xlsx")

end_time = time.time()

print(f"Time used: {end_time-start_time} seconds")
